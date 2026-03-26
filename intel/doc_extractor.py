"""
Binary document extractor — PDF, DOCX, XLSX → clean text for knowledge base indexing.

Handles:
  PDF  — pdfminer.six: preserves page structure, extracts headings from font size
  DOCX — python-docx: preserves heading hierarchy, tables → markdown tables
  XLSX — openpyxl: extracts all sheets, formats Q&A rows, skips empty cells

Usage:
  from intel.doc_extractor import extract_file, extract_pdf, extract_docx, extract_xlsx
  text = extract_file("/path/to/file.pdf")  # auto-dispatch by extension
"""

import re
import io
from pathlib import Path


# ─── PDF Extraction ───────────────────────────────────────────────────────────

def extract_pdf(path: Path, max_pages: int = 0) -> str:
    """
    Extract text from PDF preserving page structure.
    max_pages=0 means all pages.
    Returns clean text string.
    """
    try:
        from pdfminer.high_level import extract_pages
        from pdfminer.layout import LTTextContainer, LTChar, LTAnon, LTTextBox, LTTextLine, LTFigure
    except ImportError:
        return ""

    lines = []
    page_num = 0

    try:
        for page_layout in extract_pages(str(path)):
            page_num += 1
            if max_pages and page_num > max_pages:
                break

            lines.append(f"\n--- Page {page_num} ---")
            for element in page_layout:
                if isinstance(element, LTTextContainer):
                    text = element.get_text()
                    # Clean up: collapse excessive whitespace, remove form feeds
                    text = re.sub(r"\f", "\n", text)
                    text = re.sub(r"\n{3,}", "\n\n", text)
                    text = text.strip()
                    if text and len(text) > 3:
                        lines.append(text)
    except Exception as e:
        return f"[PDF extraction error: {e}]"

    return "\n".join(lines)


def extract_pdf_simple(path: Path, max_pages: int = 0) -> str:
    """Simpler PDF extraction using pdfminer high-level API."""
    try:
        from pdfminer.high_level import extract_text
        text = extract_text(str(path), maxpages=max_pages or 0)
        # Clean up
        text = re.sub(r"\f", "\n---\n", text)   # form feed = page break
        text = re.sub(r"\n{4,}", "\n\n", text)
        return text.strip()
    except ImportError:
        return "[pdfminer.six not installed]"
    except Exception as e:
        return f"[PDF extraction error: {e}]"


# ─── DOCX Extraction ──────────────────────────────────────────────────────────

def extract_docx(path: Path) -> str:
    """
    Extract text from DOCX preserving heading hierarchy.
    Headings → markdown ## / ### markers.
    Tables → pipe-separated rows.
    """
    try:
        import docx
    except ImportError:
        return "[python-docx not installed]"

    try:
        doc = docx.Document(str(path))
    except Exception as e:
        return f"[DOCX extraction error: {e}]"

    lines = []

    for element in doc.element.body:
        tag = element.tag.split("}")[-1] if "}" in element.tag else element.tag

        if tag == "p":
            # Paragraph
            para = docx.text.paragraph.Paragraph(element, doc)
            text = para.text.strip()
            if not text:
                continue
            style = para.style.name if para.style else ""
            if "Heading 1" in style or style == "Title":
                lines.append(f"\n# {text}")
            elif "Heading 2" in style:
                lines.append(f"\n## {text}")
            elif "Heading 3" in style:
                lines.append(f"\n### {text}")
            elif "Heading" in style:
                lines.append(f"\n#### {text}")
            else:
                lines.append(text)

        elif tag == "tbl":
            # Table
            table = docx.table.Table(element, doc)
            for i, row in enumerate(table.rows):
                cells = [c.text.strip().replace("\n", " ") for c in row.cells]
                # Deduplicate merged cells
                deduped = []
                prev = None
                for c in cells:
                    if c != prev:
                        deduped.append(c)
                    prev = c
                if any(deduped):
                    lines.append(" | ".join(deduped))
                    if i == 0:
                        lines.append(" | ".join(["---"] * len(deduped)))

    text = "\n".join(lines)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


# ─── XLSX Extraction ──────────────────────────────────────────────────────────

def extract_xlsx(path: Path) -> str:
    """
    Extract content from Excel sheets.
    Formats as:
      ## Sheet Name
      Header1 | Header2 | ...
      value1  | value2  | ...

    Skips sheets that look like pure trackers (all numeric, no text questions).
    Prioritises sheets with Q&A content (long text cells).
    """
    try:
        import openpyxl
    except ImportError:
        return "[openpyxl not installed]"

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as e:
        return f"[XLSX extraction error: {e}]"

    sections = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            # Skip entirely empty rows
            non_empty = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if not non_empty:
                continue
            rows.append([str(c).strip() if c is not None else "" for c in row])

        if not rows:
            continue

        # Heuristic: skip pure-tracker sheets (all cells very short, no question text)
        all_text = " ".join(" ".join(r) for r in rows)
        avg_cell_len = len(all_text) / max(sum(len(r) for r in rows), 1)
        has_questions = any(
            len(cell) > 40
            for row in rows
            for cell in row
        )
        if not has_questions and avg_cell_len < 15:
            continue  # pure status/date tracker, no useful text

        section_lines = [f"\n## {sheet_name}"]

        # Find header row (first non-empty row)
        header = rows[0]
        # Trim trailing empty headers
        while header and not header[-1]:
            header.pop()

        section_lines.append(" | ".join(header))
        section_lines.append(" | ".join(["---"] * len(header)))

        for row in rows[1:]:
            # Pad or trim to header length
            padded = (row + [""] * len(header))[:len(header)]
            # Only include rows with some non-trivial content
            content = " ".join(padded)
            if len(content.strip()) > 5:
                section_lines.append(" | ".join(padded))

        sections.append("\n".join(section_lines))

    wb.close()
    return "\n\n".join(sections)


# ─── Dispatcher ───────────────────────────────────────────────────────────────

def extract_file(path, max_pdf_pages: int = 0) -> str:
    """
    Extract text from any supported file type.
    max_pdf_pages=0 means all pages. Set e.g. 150 for large PDFs to cap cost.
    """
    path = Path(path)
    if not path.exists():
        return ""

    ext = path.suffix.lower()

    if ext == ".pdf":
        return extract_pdf_simple(path, max_pages=max_pdf_pages)
    elif ext == ".docx":
        return extract_docx(path)
    elif ext in (".xlsx", ".xls"):
        return extract_xlsx(path)
    else:
        # Try reading as plain text
        try:
            return path.read_text(encoding="utf-8", errors="replace")
        except Exception:
            return ""


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python -m intel.doc_extractor <file_path> [max_pages]")
        sys.exit(1)
    fp = Path(sys.argv[1])
    pages = int(sys.argv[2]) if len(sys.argv) > 2 else 0
    text = extract_file(fp, max_pdf_pages=pages)
    print(text[:3000])
    print(f"\n--- Total chars: {len(text)} ---")
